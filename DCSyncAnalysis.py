import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter, defaultdict
import sys
import re

def analyze_passwords(hashes_file, cracked_file, output_file, company_keyword):
    # Read cracked passwords into hash->password mapping
    cracked = {}
    with open(cracked_file, 'r', encoding='utf-8') as f:
        for line in f:
            if ':' in line:
                hash_val, password = line.strip().split(':', 1)
                if password:  # Only store if password isn't empty
                    cracked[hash_val.lower()] = password  # Convert to lowercase for consistency

    print(f"Loaded {len(cracked)} cracked hashes")

    # Read NTLM hashes from the dump file
    hash_usage = Counter()
    format_detected = False
    
    # Debug: Check first few lines of hash file
    print(f"Reading hash file: {hashes_file}")
    with open(hashes_file, 'r', encoding='utf-8', errors='ignore') as f:
        first_lines = [f.readline().strip() for _ in range(5)]
        print(f"First few lines of hash file:")
        for i, line in enumerate(first_lines):
            print(f"Line {i+1}: {line}")
    
    with open(hashes_file, 'r', encoding='utf-8', errors='ignore') as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line or line.startswith('#'):
                continue
                
            # Try multiple formats
            
            # Format 1: domain\username:rid:lmhash:ntlmhash:::
            if ':' in line and len(line.split(':')) >= 4:
                parts = line.split(':')
                ntlm_hash = parts[3].lower()  # 4th field is NTLM hash
                if ntlm_hash and ntlm_hash != 'aad3b435b51404eeaad3b435b51404ee':  # Skip empty hashes
                    hash_usage[ntlm_hash] += 1
                    if not format_detected:
                        print(f"Detected format: domain\\username:rid:lmhash:ntlmhash:::")
                        format_detected = True
            
            # Format 2: just the hash
            elif len(line.strip()) == 32:
                ntlm_hash = line.lower()
                hash_usage[ntlm_hash] += 1
                if not format_detected:
                    print(f"Detected format: hash only (32 characters)")
                    format_detected = True
            
            # Format 3: username:hash format
            elif ':' in line and len(line.split(':')) == 2:
                _, ntlm_hash = line.split(':')
                ntlm_hash = ntlm_hash.lower()
                if len(ntlm_hash) == 32:
                    hash_usage[ntlm_hash] += 1
                    if not format_detected:
                        print(f"Detected format: username:hash")
                        format_detected = True
                    
            # Add debugging for difficult lines
            if i < 10 and not hash_usage and line:
                print(f"Debug - Line {i+1} not recognized: {line}")

    # Calculate password statistics
    password_usage = defaultdict(int)
    users_cracked = 0
    matched_hashes = []
    
    # For each hash in the hash_usage, if we cracked it, add its count to the password stats
    for hash_val, count in hash_usage.items():
        if hash_val in cracked:
            matched_hashes.append(hash_val)
            password = cracked[hash_val]
            password_usage[password] += count
            users_cracked += count

    total_hashes = sum(hash_usage.values())
    
    print(f"Total unique NTLM hashes: {len(hash_usage)}")
    print(f"Total users: {total_hashes}")
    print(f"Matched {len(matched_hashes)} hashes")
    print(f"Users with cracked passwords: {users_cracked}")

    if not password_usage:
        print("No passwords were successfully matched!")
        return

    # Create Excel workbook with nice formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Apply some basic styling
    title_font = Font(size=14, bold=True)
    header_font = Font(size=12, bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Write summary statistics
    ws['A1'] = "Password Analysis Summary"
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A3'] = "Total Users"
    ws['B3'] = total_hashes
    ws['A4'] = "Unique Hashes"
    ws['B4'] = len(hash_usage)
    ws['A5'] = "Cracked Hashes"
    ws['B5'] = len(matched_hashes)
    ws['A6'] = "Users with Cracked Passwords"
    ws['B6'] = users_cracked
    ws['A7'] = "Success Rate"
    ws['B7'] = f"{(users_cracked/total_hashes)*100:.1f}%"
    
    # Format headers
    for cell in ws['A3:A7']:
        cell[0].font = header_font
        
    # Create pie chart for success rate
    pie = PieChart()
    pie.title = "Password Cracking Success Rate"
    
    # Add data to worksheet for the chart
    ws['E3'] = "Status"
    ws['F3'] = "Users"
    ws['E3'].font = header_font
    ws['F3'].font = header_font
    ws['E3'].fill = header_fill
    ws['F3'].fill = header_fill
    
    ws['E4'] = "Cracked"
    ws['F4'] = users_cracked
    ws['E5'] = "Not Cracked"
    ws['F5'] = total_hashes - users_cracked
    
    data = Reference(ws, min_col=6, min_row=4, max_row=5)
    labels = Reference(ws, min_col=5, min_row=4, max_row=5)
    pie.add_data(data)
    pie.set_categories(labels)
    
    ws.add_chart(pie, "E8")

    # Create "Most Common Passwords" sheet
    common_pw = wb.create_sheet("Common Passwords")
    
    common_pw['A1'] = "Most Common Passwords"
    common_pw['A1'].font = title_font
    common_pw.merge_cells('A1:D1')
    common_pw['A1'].alignment = Alignment(horizontal="center")
    
    common_pw['A3'] = "Password"
    common_pw['B3'] = "Number of Users"
    common_pw['C3'] = "Percentage of Cracked"
    
    for col in ['A', 'B', 'C']:
        common_pw[f'{col}3'].font = header_font
        common_pw[f'{col}3'].fill = header_fill
    
    row = 4
    for pwd, count in sorted(password_usage.items(), key=lambda x: x[1], reverse=True):
        common_pw[f'A{row}'] = pwd
        common_pw[f'B{row}'] = count
        if users_cracked > 0:
            common_pw[f'C{row}'] = f"{(count/users_cracked)*100:.1f}%"
        row += 1

    # Create bar chart for top passwords
    chart = BarChart()
    chart.title = "Top 10 Most Used Passwords"
    chart.x_axis.title = "Password"
    chart.y_axis.title = "Number of Users"
    
    data = Reference(common_pw, min_col=2, min_row=4, max_row=13, max_col=2)
    cats = Reference(common_pw, min_col=1, min_row=4, max_row=13, max_col=1)
    
    chart.add_data(data)
    chart.set_categories(cats)
    
    common_pw.add_chart(chart, "E4")

    # Create "Password Patterns" sheet
    patterns = wb.create_sheet("Password Patterns")
    patterns['A1'] = "Password Pattern Analysis"
    patterns['A1'].font = title_font
    patterns.merge_cells('A1:D1')
    patterns['A1'].alignment = Alignment(horizontal="center")
    
    # Basic patterns
    total_unique_passwords = len(password_usage)
    uppercase = sum(1 for p in password_usage.keys() if any(c.isupper() for c in p))
    lowercase = sum(1 for p in password_usage.keys() if any(c.islower() for c in p))
    numbers = sum(1 for p in password_usage.keys() if any(c.isdigit() for c in p))
    special = sum(1 for p in password_usage.keys() if any(not c.isalnum() for c in p))
    
    patterns['A3'] = "Character Type Analysis"
    patterns['A3'].font = header_font
    patterns.merge_cells('A3:B3')
    
    patterns['A4'] = "Contains Uppercase"
    patterns['B4'] = f"{uppercase} ({uppercase/total_unique_passwords*100:.1f}%)"
    patterns['A5'] = "Contains Lowercase"
    patterns['B5'] = f"{lowercase} ({lowercase/total_unique_passwords*100:.1f}%)"
    patterns['A6'] = "Contains Numbers"
    patterns['B6'] = f"{numbers} ({numbers/total_unique_passwords*100:.1f}%)"
    patterns['A7'] = "Contains Special"
    patterns['B7'] = f"{special} ({special/total_unique_passwords*100:.1f}%)"
    
    # Character type pie chart
    pie2 = PieChart()
    pie2.title = "Character Types in Passwords"
    
    # Add data for pie chart
    patterns['D4'] = "Character Type"
    patterns['E4'] = "Count"
    patterns['D4'].font = header_font
    patterns['E4'].font = header_font
    patterns['D4'].fill = header_fill
    patterns['E4'].fill = header_fill
    
    patterns['D5'] = "Uppercase"
    patterns['E5'] = uppercase
    patterns['D6'] = "Lowercase"
    patterns['E6'] = lowercase
    patterns['D7'] = "Numbers"
    patterns['E7'] = numbers
    patterns['D8'] = "Special"
    patterns['E8'] = special
    
    data = Reference(patterns, min_col=5, min_row=5, max_row=8)
    cats = Reference(patterns, min_col=4, min_row=5, max_row=8)
    pie2.add_data(data)
    pie2.set_categories(cats)
    
    patterns.add_chart(pie2, "D10")

    # Find common words in passwords
    # Extract words (3+ letter sequences) from passwords
    words = Counter()
    company_count = 0
    
    for pwd, count in password_usage.items():
        # Look for the company name specifically
        if company_keyword.lower() in pwd.lower():
            company_count += count
        
        # Extract other words (3+ letters)
        matches = re.findall(r'[a-zA-Z]{3,}', pwd)
        for match in matches:
            words[match.lower()] += count

    # Add the company keyword separately (it might be overcounted, but we want to highlight it)
    words[company_keyword.lower()] = company_count

    patterns['A9'] = "Common Words in Passwords"
    patterns['A9'].font = header_font
    patterns.merge_cells('A9:B9')
    
    patterns['A10'] = "Word"
    patterns['B10'] = "Occurrences"
    patterns['C10'] = "% of Users"
    patterns['A10'].font = header_font
    patterns['B10'].font = header_font
    patterns['C10'].font = header_font
    patterns['A10'].fill = header_fill
    patterns['B10'].fill = header_fill
    patterns['C10'].fill = header_fill
    
    row = 11
    for word, count in words.most_common(20):
        patterns[f'A{row}'] = word
        patterns[f'B{row}'] = count
        patterns[f'C{row}'] = f"{(count/users_cracked)*100:.1f}%"
        row += 1
    
    # Create bar chart for common words
    word_chart = BarChart()
    word_chart.title = "Top 10 Common Words in Passwords"
    word_chart.x_axis.title = "Word"
    word_chart.y_axis.title = "Occurrences"
    
    word_data = Reference(patterns, min_col=2, min_row=11, max_row=20, max_col=2)
    word_cats = Reference(patterns, min_col=1, min_row=11, max_row=20, max_col=1)
    
    word_chart.add_data(word_data)
    word_chart.set_categories(word_cats)
    
    patterns.add_chart(word_chart, "E20")

    # Add length analysis
    lengths = wb.create_sheet("Password Lengths")
    
    # Calculate length statistics
    length_stats = Counter()
    for pwd, count in password_usage.items():
        length_stats[len(pwd)] += count
    
    lengths['A1'] = "Password Length Analysis"
    lengths['A1'].font = title_font
    lengths.merge_cells('A1:C1')
    lengths['A1'].alignment = Alignment(horizontal="center")
    
    lengths['A3'] = "Length"
    lengths['B3'] = "Count"
    lengths['C3'] = "Percentage"
    lengths['A3'].font = header_font
    lengths['B3'].font = header_font
    lengths['C3'].font = header_font
    lengths['A3'].fill = header_fill
    lengths['B3'].fill = header_fill
    lengths['C3'].fill = header_fill
    
    row = 4
    for length, count in sorted(length_stats.items()):
        lengths[f'A{row}'] = length
        lengths[f'B{row}'] = count
        lengths[f'C{row}'] = f"{(count/users_cracked)*100:.1f}%"
        row += 1
    
    # Create bar chart for password lengths
    length_chart = BarChart()
    length_chart.title = "Password Length Distribution"
    length_chart.x_axis.title = "Length"
    length_chart.y_axis.title = "Count"
    
    length_data = Reference(lengths, min_col=2, min_row=4, max_row=row-1, max_col=2)
    length_cats = Reference(lengths, min_col=1, min_row=4, max_row=row-1, max_col=1)
    
    length_chart.add_data(length_data)
    length_chart.set_categories(length_cats)
    
    lengths.add_chart(length_chart, "E4")

    # Save workbook
    wb.save(output_file)
    
    # Print summary to console
    print(f"\nPassword Analysis Complete!")
    print(f"Total Users: {total_hashes}")
    print(f"Cracked Passwords (Unique): {len(matched_hashes)}")
    print(f"Users with Cracked Passwords: {users_cracked}")
    print(f"Success Rate: {(users_cracked/total_hashes)*100:.1f}%")
    
    if users_cracked > 0:
        print("\nTop 5 Most Used Passwords:")
        for pwd, count in sorted(password_usage.items(), key=lambda x: x[1], reverse=True)[:5]:
            print(f"'{pwd}': {count} users ({(count/users_cracked)*100:.1f}%)")
        
        print("\nTop 5 Common Words in Passwords:")
        for word, count in words.most_common(5):
            print(f"'{word}': {count} occurrences ({(count/users_cracked)*100:.1f}%)")
    
    print(f"\nReport saved to: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("Usage: python script.py hashes_file cracked_file output_file company_keyword")
        sys.exit(1)
    
    analyze_passwords(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
